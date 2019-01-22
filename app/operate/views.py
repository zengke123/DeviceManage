from . import operate
from .. import db
from ..models import Host, Capacity
from flask import render_template, request, jsonify, send_from_directory, redirect
from flask_login import login_required
from werkzeug.utils import secure_filename
from ..filesetting import UPLOAD_FOLDER, ALLOWED_EXTENSIONS, TEMPLATE_FOLDER


# 操作页面主页，页面显示数据库中所有字段
@operate.route('/new')
@login_required
def new():
    show_cols_h = [('* 平台', 'platform'), ('* 集群', 'cluster'), ('* 主机', 'hostname'),
                   ('设备类型', 'device_type'), ('设备厂家', 'manufacturer'), ('设备型号', 'device_model'),
                   ('序列号', 'serial'), ('账户', 'account'),('业务版本', 'version'),('软件模块', 'software_version'),
                   ('内网IP', 'local_ip'),('映射IP', 'nat_ip'),('操作系统', 'os_version'),
                   ('机房', 'engine_room'), ('机架', 'frame_number'), ('电源柜', 'power_frame_number'),
                   ('入网时间', 'net_time'), ('过保时间', 'period'), ('状态', 'status')]
    show_cols_s = [('* 平台', 'platform'), ('* 集群', 'cluster'), ('硬件容量(W)', 'h_capacity'),
                   ('硬件容量(CAPS)', 'h_caps'), ('软件容量(W)', 's_capacity'), ('软件容量(CAPS)', 's_caps')]
    return render_template("new_device.html", show_cols_h=show_cols_h, show_cols_s=show_cols_s)


# 新增主机设备信息
@operate.route('/add_new', methods=["GET", "POST"])
@login_required
def add_new():
    datas = request.get_json()
    try:
        add_host = Host(**datas)
        db.session.add(add_host)
        db.session.commit()
        result = {"flag": "success"}
    except Exception as e:
        print(str(e))
        result = {"flag": "fail"}
    return jsonify(result)


# 新增软件容量信息
@operate.route('/add_new_s', methods=["GET", "POST"])
@login_required
def add_new_s():
    datas = request.get_json()
    try:
        add_host = Capacity(**datas)
        db.session.add(add_host)
        db.session.commit()
        result = {"flag": "success"}
    except Exception as e:
        print(str(e))
        result = {"flag": "fail"}
    return jsonify(result)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@operate.route('/load', methods=['GET', 'POST'])
@login_required
def load():
    import os
    errors = ""
    if request.method == 'POST':
        # 请求中无文件
        if 'file' not in request.files:
            return redirect(request.url)
        file = request.files['file']
        # 文件名为空
        if file.filename == '':
            return redirect(request.url)
        # 文件符合要求
        if file and allowed_file(file.filename):
            # fieldvalue为前端页面传入的select中option选中参数，该参数定义为数据的类型
            data_type = request.form.get("fieldvalue")
            filename = secure_filename(file.filename)
            file.save(os.path.join(UPLOAD_FOLDER, filename))
            # 调用入库函数
            nums = load_to_db(data_type, UPLOAD_FOLDER+filename)
            return  render_template("load_success.html", nums=nums)
        errors = '上传文件失败，只支持xlsx格式'
    return render_template("load_device.html",errors=errors)



@operate.route('/download_file/<filename>')
@login_required
def download_file(filename):
    # path = "/Users/EB/PycharmProjects/DeviceManage/unloadfiles/"
    # file = request.form.get("filename")
    return send_from_directory(TEMPLATE_FOLDER, filename=filename, as_attachment=True)


def load_to_db(data_type, filename):
    from openpyxl import load_workbook
    nums = 0
    # 与数据中字段一致
    host_names = ["platform", "cluster", "hostname", "device_type", "manufacturer", "device_model", "serial",
             "account", "version", "software_version", "local_ip", "nat_ip", "os_version", "engine_room",
             "frame_number", "power_frame_number", "net_time", "period", "status"]
    capacity_names = ["platform", "cluster", "h_capacity", "h_caps", "s_capacity", "s_caps"]
    infos = {
        "hardware": {
            "cols": host_names,
            "sheet": "设备列表"
        },
        "software": {
            "cols": capacity_names,
            "sheet": "系统容量"
        }
    }
    try:
        wb = load_workbook(filename)
        # 读取导入excel文件中的sheet表
        ws = wb[infos.get(data_type)["sheet"]]
        # 获取列数
        cols = len(infos.get(data_type)["cols"])
        # 获取行数
        rows = ws.max_row
        dbs = []
        # 读取excel中每行数据，构造成dict, 方便入库
        for i in range(2, rows + 1):
            data = []
            for j in range(1, cols + 1):
                value = ws.cell(i, j).value
                data.append(value)
            item = {k: v for k, v in zip(infos.get(data_type)["cols"], data)}
            if data_type == "hardware":
                dbs.append(Host(**item))
            elif data_type == "software":
                dbs.append(Capacity(**item))
        # 批量添加
        db.session.add_all(dbs)
        db.session.commit()
        nums = len(dbs)
    except Exception as e:
        print(str(e))
    finally:
        # 最后删除上传的文件
        import os
        os.remove(filename)
    return nums
